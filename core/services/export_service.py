"""
Export service - moved from core/services.py for better organization.
"""

import csv
import io
from datetime import datetime
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill


class ExportService:
    """
    Server-side data export service.
    Handles CSV and Excel exports with proper formatting.
    """
    
    @staticmethod
    def export_members_csv(members, church_name=None):
        """Export members to CSV format."""
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Headers
        headers = [
            'Member ID', 'First Name', 'Last Name', 'Email', 'Phone',
            'Gender', 'Date of Birth', 'Membership Date', 'Status',
            'Address', 'Occupation', 'Place of Work'
        ]
        writer.writerow(headers)
        
        # Member data
        for member in members:
            writer.writerow([
                member.member_id,
                member.first_name,
                member.last_name,
                member.email,
                member.phone,
                member.gender,
                member.date_of_birth.strftime('%Y-%m-%d') if member.date_of_birth else '',
                member.membership_date.strftime('%Y-%m-%d') if member.membership_date else '',
                member.status,
                member.address,
                member.profession or member.occupational_status or '',
                member.place_of_work or ''
            ])
        
        output.seek(0)
        response = HttpResponse(output.getvalue(), content_type='text/csv')
        filename = f"{church_name or 'Church'}_Members_{datetime.now().strftime('%Y%m%d')}.csv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
    
    @staticmethod
    def export_members_excel(members, church_name=None):
        """Export members to Excel format."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Members"
        
        # Headers with styling
        headers = [
            'Member ID', 'First Name', 'Last Name', 'Email', 'Phone',
            'Gender', 'Date of Birth', 'Membership Date', 'Status',
            'Address', 'Occupation', 'Place of Work'
        ]
        
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Member data
        for row, member in enumerate(members, 2):
            ws.cell(row=row, column=1).value = member.member_id
            ws.cell(row=row, column=2).value = member.first_name
            ws.cell(row=row, column=3).value = member.last_name
            ws.cell(row=row, column=4).value = member.email
            ws.cell(row=row, column=5).value = member.phone
            ws.cell(row=row, column=6).value = member.gender
            ws.cell(row=row, column=7).value = member.date_of_birth.strftime('%Y-%m-%d') if member.date_of_birth else ''
            ws.cell(row=row, column=8).value = member.membership_date.strftime('%Y-%m-%d') if member.membership_date else ''
            ws.cell(row=row, column=9).value = member.status
            ws.cell(row=row, column=10).value = member.address
            ws.cell(row=row, column=11).value = member.profession or member.occupational_status or ''
            ws.cell(row=row, column=12).value = member.place_of_work or ''
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add metadata sheet
        ws_meta = wb.create_sheet(title="Export Info")
        ws_meta['A1'] = 'Church Name'
        ws_meta['B1'] = church_name or 'Unknown Church'
        ws_meta['A2'] = 'Export Date'
        ws_meta['B2'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws_meta['A3'] = 'Total Members'
        ws_meta['B3'] = len(list(members))
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"{church_name or 'Church'}_Members_{datetime.now().strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
    
    @staticmethod
    def export_events_csv(events, church_name=None):
        """Export events to CSV."""
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Headers
        headers = [
            'Title', 'Description', 'Date', 'End Date', 'Location',
            'Type', 'Capacity', 'Attendees', 'Recurring', 'Pattern'
        ]
        writer.writerow(headers)
        
        # Event data
        for event in events:
            writer.writerow([
                event.title,
                event.description,
                event.date.strftime('%Y-%m-%d %H:%M') if event.date else '',
                event.end_date.strftime('%Y-%m-%d %H:%M') if event.end_date else '',
                event.location,
                event.type,
                event.capacity or '',
                len(event.attendees) if event.attendees else 0,
                'Yes' if event.is_recurring else 'No',
                event.recurrence_pattern or ''
            ])
        
        output.seek(0)
        response = HttpResponse(output.getvalue(), content_type='text/csv')
        filename = f"{church_name or 'Church'}_Events_{datetime.now().strftime('%Y%m%d')}.csv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
    
    @staticmethod
    def export_payments_csv(payments, church_name=None, year=None):
        """Export payments to CSV."""
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Headers
        headers = [
            'Date', 'Member', 'Type', 'Amount', 'Currency', 'Method',
            'Reference', 'Status', 'Receipt Number', 'Notes'
        ]
        writer.writerow(headers)
        
        # Payment data
        total_amount = 0
        for payment in payments:
            writer.writerow([
                payment.date.strftime('%Y-%m-%d') if payment.date else '',
                payment.member.full_name if payment.member else '',
                payment.type,
                payment.amount,
                payment.currency,
                payment.method,
                payment.reference,
                payment.status,
                payment.receipt_number or '',
                payment.notes or ''
            ])
            if payment.status == 'completed':
                total_amount += payment.amount
        
        # Add total row
        writer.writerow([])
        writer.writerow(['TOTAL', '', '', total_amount, '', '', '', '', '', ''])
        
        output.seek(0)
        response = HttpResponse(output.getvalue(), content_type='text/csv')
        year_str = f"_{year}" if year else ''
        filename = f"{church_name or 'Church'}_Payments{year_str}_{datetime.now().strftime('%Y%m%d')}.csv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response

